from datetime import date
from tkinter import *
from tkinter import ttk
import tkinter as tk
from tkcalendar import *
from openpyxl import Workbook
import pathlib
import openpyxl
import xlrd
from tkinter import messagebox

sp = Tk()
#title and icon set

sp.title("My Spend")
sp.iconbitmap('myspend.ico')

#excel file Create 
file = pathlib.Path("backend_Data.xlsx")

if file.exists ():

    pass

else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]= "SPEND TYPE"
    sheet["C1"]= "DATE"
    sheet["B1"]= "AMOUNT"
    

file.save("backend_Data.xlsx")
#submit button function and save data in excel
def myfun():
    file = openpyxl.load_workbook("backend_Data.xlsx")
    sheet=file.active
    sheet.cell(column=1, row=sheet.max_row+1,value=clicked.get())
    sheet.cell(column=2, row=sheet.max_row,value=a2.get())
    sheet.cell(column=3, row=sheet.max_row,value=d2.get())
    file.save("backend_Data.xlsx")
    messagebox.showinfo('Confirmation','Data Saved Successfully')

#windoes size 
sp.geometry("400x450")
sp.maxsize(400, 450)
sp.minsize(400, 450)

#1st label - My Mounthly Spends
l2 = Label(sp, text="My Mounthly Spends" , font = ('Time',20,'bold'))
l2.place(x = 60, y = 30 )

#2nd label - spend type
l1 = Label(sp, text="Spend Type : " , font = "time 10 ")
l1.place(x = 60, y = 100 )

#drop down boxes
spendtype=["light","Gass Bill","Phone - Aai ","Phone - Sukanya ","Phone - Sanket","Phone - Baba ","Phone - Mavshi"]
clicked=ttk.Combobox(sp,value=spendtype,width=20)
clicked.place(x = 150 , y = 100)
clicked.current(0)

#amount box and lebel
a1 = Label(sp,text="Amount :", font = "time 10")
a1.place(x = 60, y = 150 )
a2 = Entry(sp,width = 15, bd = 3)
a2.place(x=150,y =150 )

#date lebel and date box
d1 = Label(sp,text="Date :", font = "time 10")
d1.place(x = 60, y = 200 )
d2 = DateEntry(sp,selectmode='day')
d2.place(x=150,y = 200 )

#submite button
Button(sp,text="Submit" ,bg="black",command =myfun, fg="white",font="Time",width=10).place(x=150 , y=250)

#data dsave popup
sp.mainloop()
